"""Populate the TPR Excel template with country data, or build from scratch."""
from __future__ import annotations

import io
import re
import zipfile
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from loguru import logger

if TYPE_CHECKING:
    from src.viz.profile import CountryProfile

# ── Sheet configuration ───────────────────────────────────────────────────────
SHEET_CONFIG: list[tuple[str, str, list[str]]] = [
    ("patent_applications",        "Patent Applications",              ["year", "resident", "non_resident", "total"]),
    ("patent_grants",              "Patent Grants",                    ["year", "resident", "non_resident", "total"]),
    ("trademark_applications",     "Trademark Applications",           ["year", "resident", "non_resident", "total"]),
    ("trademark_registrations",    "Trademark Registrations",          ["year", "resident", "non_resident", "total"]),
    ("design_applications",        "Industrial Design Applications",   ["year", "resident", "non_resident", "total"]),
    ("design_registrations",       "Industrial Design Registrations",  ["year", "resident", "non_resident", "total"]),
    ("utility_model_applications", "Utility Model Applications",       ["year", "resident", "non_resident", "total"]),
    ("utility_model_grants",       "Utility Model Grants",             ["year", "resident", "non_resident", "total"]),
    ("geographical_indications",   "Geographical Indications",         ["year", "total"]),
    ("ip_services",                "(BOP) Charges for the Use of IP",  ["year", "imports_usd", "exports_usd"]),
]

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill("solid", fgColor="005A8C")
_HEADER_ALIGN = Alignment(horizontal="center")

_SUMMARY_ROW_MAP: list[tuple[int, str, str]] = [
    (2,  "Patent Applications",                    "WIPO"),
    (3,  "Patent Grants",                           "WIPO"),
    (4,  "Trademark Applications",                  "WIPO"),
    (5,  "Trademark Registrations",                 "WIPO"),
    (6,  "Industrial Design Applications",          "WIPO"),
    (7,  "Industrial Design Registrations",         "WIPO"),
    (8,  "Utility Model Applications",              "WIPO"),
    (9,  "Utility Model Grants",                    "WIPO"),
    (10, "Geographical Indications",                "WIPO"),
    (11, "Charges for the Use of IP (USD million)", "WTO"),
]
_WIPO_SOURCE = "WIPO IP Statistics Data Center. Viewed at: https://www3.wipo.int/ipstats ({date})."
_WTO_SOURCE  = "WTO Stats Portal. Viewed at: https://stats.wto.org/ ({date})."

_YEARS_TO_WRITE = 7

# Map data sheet name → Summary label (for chart titles)
_SHEET_TO_LABEL: dict[str, str] = {
    cfg[1]: row[1]
    for cfg, row in zip(SHEET_CONFIG, _SUMMARY_ROW_MAP)
}


# ── Public entry point ────────────────────────────────────────────────────────

def write_country_workbook(
    profile: "CountryProfile",
    template_path: Path | None = None,
) -> io.BytesIO:
    tp = template_path or Path("templates") / "TPR_IP_Template.xlsx"
    if tp.exists():
        logger.info(f"Populating template via zip: {tp}")
        return _populate_template_zip(tp, profile)
    logger.warning("Template not found — building workbook from scratch")
    return _to_buffer(_build_scratch(profile))


# ── Zip-based template population (preserves all chart XML) ──────────────────

def _populate_template_zip(tp: Path, profile: "CountryProfile") -> io.BytesIO:
    """Surgically update cell data inside the xlsx zip, leaving charts intact."""
    pulled   = date.today().strftime("%d/%m/%Y")
    data_map = _build_data_map(profile)

    # Build filtered data per sheet and compute actual year ranges
    sheet_data: dict[str, tuple[pd.DataFrame, list[str]]] = {}
    sheet_year_range: dict[str, str] = {}
    for attr, hint, columns in SHEET_CONFIG:
        df = _latest_years(data_map.get(attr, pd.DataFrame()), _YEARS_TO_WRITE)
        sheet_data[hint] = (df, columns)
        if not df.empty and "year" in df.columns:
            sheet_year_range[hint] = f"{int(df['year'].min())}-{int(df['year'].max())}"
        else:
            display_start = profile.end_year - (_YEARS_TO_WRITE - 1)
            sheet_year_range[hint] = f"{display_start}-{profile.end_year}"

    # Per-indicator titles using actual year ranges
    new_titles = [
        f"{profile.country_name} {label}: {sheet_year_range[hint]}"
        for (_, label, _), (_, hint, _) in zip(_SUMMARY_ROW_MAP, SHEET_CONFIG)
    ]
    wipo_src = _WIPO_SOURCE.format(date=pulled)
    wto_src  = _WTO_SOURCE.format(date=pulled)

    template_bytes = tp.read_bytes()

    with zipfile.ZipFile(io.BytesIO(template_bytes), "r") as zin:
        sheet_file_map = _build_sheet_file_map(zin)

        # Update shared strings (string-based, no ET serialisation)
        ss_xml    = zin.read("xl/sharedStrings.xml")
        new_ss, d_indices, wipo_idx, wto_idx = _update_shared_strings(
            ss_xml, new_titles, wipo_src, wto_src
        )

        summary_file = sheet_file_map.get("Summary")

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                fname = item.filename
                data  = zin.read(fname)

                if fname == "xl/sharedStrings.xml":
                    data = new_ss

                elif summary_file and fname == f"xl/{summary_file}":
                    data = _update_summary_sheet(data, d_indices, wipo_idx, wto_idx)

                else:
                    for sheet_name, ws_file in sheet_file_map.items():
                        if fname == f"xl/{ws_file}" and sheet_name in sheet_data:
                            df, cols = sheet_data[sheet_name]
                            data = _update_data_sheet(data, df, cols)
                            break

                zout.writestr(item, data)

    output.seek(0)
    return output


def _build_sheet_file_map(zin: zipfile.ZipFile) -> dict[str, str]:
    """Return {sheet_name: 'worksheets/sheetN.xml'} using workbook.xml + rels."""
    NS  = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NSR = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    wb_root  = ET.fromstring(zin.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))

    rid_to_file: dict[str, str] = {r.get("Id", ""): r.get("Target", "") for r in rel_root}

    result: dict[str, str] = {}
    for el in wb_root.iter():
        if el.tag.endswith("}sheet"):
            name = el.get("name", "")
            rid  = el.get(f"{{{NSR}}}id", "")
            target = rid_to_file.get(rid, "")
            if target:
                result[name] = target
    return result


# ── Shared-strings update (string surgery, no ET serialisation) ──────────────

def _update_shared_strings(
    ss_xml: bytes,
    new_titles: list[str],
    wipo_src: str,
    wto_src: str,
) -> tuple[bytes, list[int], int, int]:
    """
    Parse sharedStrings.xml with ET (read-only), then do targeted string
    replacements so the output bytes preserve the original namespace declarations.
    Returns (new_xml_bytes, d_col_indices, wipo_str_idx, wto_str_idx).
    """
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root     = ET.fromstring(ss_xml)
    si_list  = root.findall(f"{{{NS}}}si")
    strings  = [_get_si_text(si, NS) for si in si_list]

    indicator_keywords = [label.split()[0].lower() for _, label, _ in _SUMMARY_ROW_MAP]

    # Find existing indices for title strings (match by keyword + year pattern)
    title_indices: list[int | None] = []
    for kw in indicator_keywords:
        idx = next(
            (i for i, s in enumerate(strings)
             if s and kw in s.lower() and re.search(r"\d{4}-\d{4}", s)),
            None,
        )
        title_indices.append(idx)

    wipo_idx_found = next((i for i, s in enumerate(strings) if s and "wipo.int" in s.lower()), None)
    wto_idx_found  = next((i for i, s in enumerate(strings) if s and "stats.wto.org" in s.lower()), None)

    xml_str = ss_xml.decode("utf-8")

    def _replace_si_text(xml: str, idx: int, new_text: str) -> str:
        """Replace the text inside the idx-th <si> element."""
        # Find the idx-th <si> occurrence
        pos = 0
        for _ in range(idx + 1):
            pos = xml.index("<si>", pos) + 4
        si_end = xml.index("</si>", pos)
        old_si_content = xml[pos:si_end]
        new_si_content = f"<t>{_xml_escape(new_text)}</t>"
        return xml[:pos] + new_si_content + xml[si_end:]

    def _append_si(xml: str, new_text: str) -> tuple[str, int]:
        """Append a new <si> before </sst> and return (new_xml, new_index)."""
        new_si = f"<si><t>{_xml_escape(new_text)}</t></si>"
        insert_pos = xml.rfind("</sst>")
        xml = xml[:insert_pos] + new_si + xml[insert_pos:]
        # Count total <si> before the new one
        idx = xml[:insert_pos].count("<si>")
        return xml, idx

    d_indices: list[int] = []
    for i, title in enumerate(new_titles):
        found_idx = title_indices[i] if i < len(title_indices) else None
        if found_idx is not None:
            xml_str = _replace_si_text(xml_str, found_idx, title)
            d_indices.append(found_idx)
        else:
            xml_str, new_idx = _append_si(xml_str, title)
            d_indices.append(new_idx)

    if wipo_idx_found is not None:
        xml_str = _replace_si_text(xml_str, wipo_idx_found, wipo_src)
        wipo_idx = wipo_idx_found
    else:
        xml_str, wipo_idx = _append_si(xml_str, wipo_src)

    if wto_idx_found is not None:
        xml_str = _replace_si_text(xml_str, wto_idx_found, wto_src)
        wto_idx = wto_idx_found
    else:
        xml_str, wto_idx = _append_si(xml_str, wto_src)

    # Update count/uniqueCount attributes
    n_si = xml_str.count("<si>")
    xml_str = re.sub(r'count="\d+"', f'count="{n_si}"', xml_str)
    xml_str = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{n_si}"', xml_str)

    return xml_str.encode("utf-8"), d_indices, wipo_idx, wto_idx


def _get_si_text(si: ET.Element, NS: str) -> str | None:
    t = si.find(f"{{{NS}}}t")
    if t is not None:
        return t.text or ""
    parts = [r.find(f"{{{NS}}}t") for r in si.findall(f"{{{NS}}}r")]
    return "".join(p.text or "" for p in parts if p is not None) or None


def _xml_escape(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


# ── Summary sheet cell update (string surgery) ───────────────────────────────

def _update_summary_sheet(
    ws_xml: bytes,
    d_indices: list[int],
    wipo_idx: int,
    wto_idx: int,
) -> bytes:
    """Replace shared-string indices in D2:D11 and E2:E11 cells."""
    xml = ws_xml.decode("utf-8")

    source_map = {r: wipo_idx for r in range(2, 11)}
    source_map[11] = wto_idx

    for row_num in range(2, 12):
        d_idx = row_num - 2
        if d_idx < len(d_indices):
            xml = _replace_cell_v(xml, f"D{row_num}", str(d_indices[d_idx]))
        xml = _replace_cell_v(xml, f"E{row_num}", str(source_map[row_num]))

    return xml.encode("utf-8")


def _replace_cell_v(xml: str, cell_ref: str, new_value: str) -> str:
    """Replace the <v> content of the cell with the given reference."""
    pattern = (
        rf'(<c r="{re.escape(cell_ref)}"[^>]*>(?:<f>[^<]*</f>)?<v>)'
        rf'[^<]*'
        rf'(</v>)'
    )
    replacement = rf'\g<1>{new_value}\g<2>'
    result = re.sub(pattern, replacement, xml)
    if result == xml:
        # Cell may have style attribute before type — try broader match
        pattern2 = (
            rf'(<c r="{re.escape(cell_ref)}"(?:[^/]|/(?!>))*?><v>)'
            rf'[^<]*'
            rf'(</v>)'
        )
        result = re.sub(pattern2, replacement, xml, flags=re.DOTALL)
    return result


# ── Data sheet row replacement (string surgery) ───────────────────────────────

def _update_data_sheet(ws_xml: bytes, df: pd.DataFrame, columns: list[str]) -> bytes:
    """Replace rows 2+ in a worksheet by surgically swapping the <sheetData> block."""
    xml = ws_xml.decode("utf-8")

    available  = [c for c in columns if not df.empty and c in df.columns]
    col_letters = [chr(ord("A") + i) for i in range(len(available))]
    max_col    = len(available)

    # Build new row XML for all data rows
    rows_xml = ""
    for row_offset, (_, row_data) in enumerate(df[available].iterrows()):
        row_num = row_offset + 2
        cells   = ""
        for col_letter, col_name in zip(col_letters, available):
            val = row_data[col_name]
            if pd.isna(val):
                continue
            num = int(val) if isinstance(val, float) and val == int(val) else val
            cells += f'<c r="{col_letter}{row_num}"><v>{num}</v></c>'
        rows_xml += f'<row r="{row_num}" spans="1:{max_col}">{cells}</row>'

    # Replace everything between the header row (r="1") and </sheetData>
    # Strategy: keep row r="1", remove all subsequent rows, insert new rows
    xml = re.sub(
        r'(<row r="1"[^>]*>.*?</row>)(.*?)(</sheetData>)',
        lambda m: m.group(1) + rows_xml + m.group(3),
        xml,
        flags=re.DOTALL,
    )

    # Update <dimension ref="...">
    if not df.empty and col_letters:
        last_row = 1 + len(df)
        xml = re.sub(
            r'(<dimension ref=")[^"]*(")',
            rf'\g<1>A1:{col_letters[-1]}{last_row}\g<2>',
            xml,
        )

    return xml.encode("utf-8")


# ── Chart title injection (string surgery) ────────────────────────────────────

def _update_chart_title(
    chart_xml: bytes,
    country_name: str,
    sheet_year_range: dict[str, str],
) -> bytes:
    """Inject title text into <c:title> by string replacement, not ET serialisation."""
    xml = chart_xml.decode("utf-8")

    # Determine which sheet this chart belongs to
    m = re.search(r"'([^']+)'!\$", xml)
    sheet_name = m.group(1) if m else None
    label = _SHEET_TO_LABEL.get(sheet_name, sheet_name) if sheet_name else None
    if not label:
        return chart_xml

    yr = sheet_year_range.get(sheet_name, "")
    title_text = f"{country_name} {label}: {yr}" if yr else f"{country_name} {label}"
    escaped    = _xml_escape(title_text)

    run_xml = f"<a:r><a:t>{escaped}</a:t></a:r>"

    # Find the title <a:p> block (inside <c:title>) and insert the run
    # before <a:endParaRPr
    def _inject_run(match: re.Match) -> str:
        p_content = match.group(0)
        # Remove any existing <a:r>...</a:r> runs
        p_content = re.sub(r"<a:r>.*?</a:r>", "", p_content, flags=re.DOTALL)
        # Insert run before <a:endParaRPr
        p_content = p_content.replace("<a:endParaRPr", run_xml + "<a:endParaRPr", 1)
        return p_content

    # Only operate on the first <a:p> within <c:title>
    title_block_m = re.search(r"<c:title>.*?</c:title>", xml, re.DOTALL)
    if not title_block_m:
        return chart_xml

    title_block = title_block_m.group(0)
    updated_title = re.sub(r"<a:p>.*?</a:p>", _inject_run, title_block, count=1, flags=re.DOTALL)
    xml = xml[:title_block_m.start()] + updated_title + xml[title_block_m.end():]

    return xml.encode("utf-8")


# ── Scratch workbook (no template) ───────────────────────────────────────────

def _build_scratch(profile: "CountryProfile") -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    data_map = _build_data_map(profile)
    for (attr, hint, columns) in SHEET_CONFIG:
        ws = wb.create_sheet(title=hint)
        df = data_map.get(attr, pd.DataFrame())
        available = [c for c in columns if not df.empty and c in df.columns] or columns

        for col_idx, col_name in enumerate(available, start=1):
            cell = ws.cell(row=1, column=col_idx, value=_header_label(col_name))
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        if not df.empty:
            _write_rows_openpyxl(ws, df, available, start_row=2)

        for col_idx in range(1, len(available) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    return wb


def _header_label(col_name: str) -> str:
    return col_name.replace("_", " ").title().replace("Usd", "USD")


# ── Shared helpers ────────────────────────────────────────────────────────────

def _latest_years(df: pd.DataFrame, n: int) -> pd.DataFrame:
    if df.empty or "year" not in df.columns:
        return df
    df = df.sort_values("year")
    value_cols = [c for c in df.columns if c != "year"]
    if value_cols:
        df = df[df[value_cols].notna().any(axis=1)]
    return df.tail(n).reset_index(drop=True)


def _write_rows_openpyxl(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    df: pd.DataFrame,
    columns: list[str],
    start_row: int = 2,
) -> None:
    available = [c for c in columns if c in df.columns]
    for row_offset, (_, row_data) in enumerate(df[available].iterrows()):
        for col_idx, col_name in enumerate(available, start=1):
            val = row_data[col_name]
            ws.cell(row=start_row + row_offset, column=col_idx,
                    value=None if pd.isna(val) else val)


def _build_data_map(profile: "CountryProfile") -> dict[str, pd.DataFrame]:
    return {
        "patent_applications":        getattr(profile, "patent_applications",        pd.DataFrame()),
        "patent_grants":              getattr(profile, "patent_grants",              pd.DataFrame()),
        "trademark_applications":     getattr(profile, "trademark_applications",     pd.DataFrame()),
        "trademark_registrations":    getattr(profile, "trademark_registrations",    pd.DataFrame()),
        "design_applications":        getattr(profile, "design_applications",        pd.DataFrame()),
        "design_registrations":       getattr(profile, "design_registrations",       pd.DataFrame()),
        "utility_model_applications": getattr(profile, "utility_model_applications", pd.DataFrame()),
        "utility_model_grants":       getattr(profile, "utility_model_grants",       pd.DataFrame()),
        "geographical_indications":   getattr(profile, "geographical_indications",   pd.DataFrame()),
        "ip_services":                getattr(profile, "ip_services",                pd.DataFrame()),
    }


def _to_buffer(wb: openpyxl.Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
